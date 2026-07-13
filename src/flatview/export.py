from __future__ import annotations

import csv
import os
import unicodedata
from pathlib import Path

from flatview.analytics import compute_stats, price_per_m2
from flatview.models import Listing


def _strip_diacritics(s: str) -> str:
    """Strip diacritics for PDF rendering (Helvetica doesn't support them)."""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _location_str(listing: Listing) -> str:
    loc = listing.city
    if listing.postcode:
        loc += f" ({listing.postcode})"
    return loc


def _summary_rows(listings: list[Listing]) -> list[list]:
    """Generate summary stat rows for export with full percentile breakdown."""
    stats = compute_stats(listings)
    price = stats.get("price") or {}
    pm2 = stats.get("pm2") or {}

    def _r(v):
        return round(v) if isinstance(v, (int, float)) else ""

    rows: list[list] = [[], ["Summary", "", "", "", "Price", "", "EUR/m2"]]
    for label, key in [
        ("Count", "n"),
        ("Min", "min"),
        ("P10", "p10"),
        ("P25", "p25"),
        ("Median (P50)", "p50"),
        ("Average", "avg"),
        ("P75", "p75"),
        ("P90", "p90"),
        ("Max", "max"),
    ]:
        if price.get(key) is None and pm2.get(key) is None:
            continue
        rows.append([label, "", "", "", _r(price.get(key)), "", _r(pm2.get(key))])
    n_outliers = sum(1 for l in listings if l.is_outlier)
    if n_outliers:
        n_bargain = sum(1 for l in listings if l.outlier_side == "bargain")
        n_over = sum(1 for l in listings if l.outlier_side == "overpriced")
        rows.append([])
        rows.append(
            [
                f"Outliers flagged (EUR/m2 IQR): {n_outliers} "
                f"({n_bargain} bargain, {n_over} overpriced)"
            ]
        )
    return rows


HEADERS = [
    "#",
    "Source",
    "Segment",
    "Title",
    "Price (EUR)",
    "Area (m2)",
    "EUR/m2",
    "Location",
    "Date",
    "URL",
]


def _listing_row(i: int, l: Listing) -> list:
    pm2 = price_per_m2(l)
    segment: str = l.segment if l.segment != "unknown" else ""
    if l.is_outlier:
        marker = f"*{l.outlier_side}" if l.outlier_side else "*"
        segment = f"{segment} {marker}" if segment else marker
    return [
        i,
        l.source,
        segment,
        l.title,
        l.price,
        l.area if l.area else "",
        round(pm2, 2) if pm2 else "",
        _location_str(l),
        l.date,
        l.url or "",
    ]


def export_csv(listings: list[Listing], path: str | Path) -> None:
    """Export listings to CSV with summary stats."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for i, l in enumerate(listings, 1):
            w.writerow(_listing_row(i, l))
        for row in _summary_rows(listings):
            w.writerow(row)


def export_xlsx(listings: list[Listing], path: str | Path) -> None:
    """Export listings to Excel with formatting and summary stats."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, l in enumerate(listings, 1):
        ws.append(_listing_row(i, l))

    # Summary
    row_num = len(listings) + 3
    for srow in _summary_rows(listings):
        if not srow:
            continue
        for col_idx, val in enumerate(srow, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if col_idx == 1:
                cell.font = Font(bold=True)
        row_num += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)


def export_pdf(listings: list[Listing], path: str | Path, title: str = "Listings") -> None:
    """Export listings to landscape A4 PDF."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, _strip_diacritics(title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    col_widths = [8, 18, 14, 70, 22, 16, 20, 32, 20, 47]
    col_headers = [
        "#",
        "Source",
        "Seg",
        "Title",
        "Price (EUR)",
        "Area",
        "EUR/m2",
        "Location",
        "Date",
        "URL",
    ]

    pdf.set_font("Helvetica", "B", 7)
    for w, h in zip(col_widths, col_headers, strict=True):
        pdf.cell(w, 5, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 6)
    for i, l in enumerate(listings, 1):
        pm2 = price_per_m2(l)
        seg: str = l.segment if l.segment != "unknown" else ""
        if l.is_outlier:
            seg = f"{seg}*" if seg else "*"
        vals = [
            str(i),
            l.source[:6],
            seg[:6],
            l.title[:50],
            f"{l.price:,.0f}" if l.price else "",
            f"{l.area:.0f}" if l.area else "",
            f"{pm2:,.0f}" if pm2 else "",
            _location_str(l),
            l.date,
            (l.url or "")[:45],
        ]
        vals = [_strip_diacritics(v) for v in vals]
        for w, v in zip(col_widths, vals, strict=True):
            pdf.cell(w, 4, v, border=1)
        pdf.ln()

    # Summary
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 6, "Summary", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 8)

    stats = compute_stats(listings)
    price = stats.get("price") or {}
    pm2_stats = stats.get("pm2") or {}

    def fmt(v):
        return f"{v:,.0f}" if isinstance(v, (int, float)) else "-"

    pdf.cell(
        0,
        5,
        f"Listings with price: {price.get('n', 0)}  |  Listings with area: {pm2_stats.get('n', 0)}",
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    if price.get("n"):
        pdf.cell(
            0,
            5,
            f"Price - P10 {fmt(price.get('p10'))}  P25 {fmt(price.get('p25'))}  "
            f"P50 {fmt(price.get('p50'))}  P75 {fmt(price.get('p75'))}  "
            f"P90 {fmt(price.get('p90'))}  avg {fmt(price.get('avg'))}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
    if pm2_stats.get("n"):
        pdf.cell(
            0,
            5,
            f"EUR/m2 - P10 {fmt(pm2_stats.get('p10'))}  P25 {fmt(pm2_stats.get('p25'))}  "
            f"P50 {fmt(pm2_stats.get('p50'))}  P75 {fmt(pm2_stats.get('p75'))}  "
            f"P90 {fmt(pm2_stats.get('p90'))}  avg {fmt(pm2_stats.get('avg'))}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
    n_outliers = sum(1 for l in listings if l.is_outlier)
    if n_outliers:
        n_bargain = sum(1 for l in listings if l.outlier_side == "bargain")
        n_over = sum(1 for l in listings if l.outlier_side == "overpriced")
        pdf.cell(
            0,
            5,
            f"Outliers flagged (EUR/m2 IQR): {n_outliers} "
            f"({n_bargain} bargain, {n_over} overpriced)",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )

    pdf.output(path)
