from __future__ import annotations

import csv
import os
import unicodedata
from pathlib import Path
from statistics import median

from b_scrape.models import Listing


def _strip_diacritics(s: str) -> str:
    """Strip diacritics for PDF rendering (Helvetica doesn't support them)."""
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")


def _price_per_m2(listing: Listing) -> float | None:
    if listing.price and listing.area and listing.area > 0:
        return listing.price / listing.area
    return None


def _location_str(listing: Listing) -> str:
    loc = listing.city
    if listing.postcode:
        loc += f" ({listing.postcode})"
    return loc


def _summary_rows(listings: list[Listing]) -> list[list]:
    """Generate summary stat rows for export."""
    prices = [l.price for l in listings if l.price is not None]
    pm2s = [_price_per_m2(l) for l in listings]
    pm2s = [v for v in pm2s if v is not None]

    rows = [[], ["Summary", "", "", "Price", "", "EUR/m2"]]
    for label, values, m2_values in [
        ("Average", prices, pm2s),
        ("Median", prices, pm2s),
        ("Min", prices, pm2s),
        ("Max", prices, pm2s),
    ]:
        if not values:
            continue
        fn = {"Average": lambda v: sum(v) / len(v), "Median": median, "Min": min, "Max": max}[label]
        p = round(fn(values))
        m = round(fn(m2_values)) if m2_values else ""
        rows.append([label, "", "", p, "", m])
    return rows


HEADERS = ["#", "Source", "Title", "Price (EUR)", "Area (m2)", "EUR/m2", "Location", "Date", "URL"]


def _listing_row(i: int, l: Listing) -> list:
    pm2 = _price_per_m2(l)
    return [
        i,
        l.source,
        l.title,
        l.price,
        l.area if l.area else "",
        round(pm2, 2) if pm2 else "",
        _location_str(l),
        l.date,
        l.url or "",
    ]


def export_csv(listings: list[Listing], path: str | Path) -> None:
    """Export listings to CSV with summary stats."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for i, l in enumerate(listings, 1):
            w.writerow(_listing_row(i, l))
        for row in _summary_rows(listings):
            w.writerow(row)


def export_xlsx(listings: list[Listing], path: str | Path) -> None:
    """Export listings to Excel with formatting and summary stats."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, l in enumerate(listings, 1):
        ws.append(_listing_row(i, l))

    # Summary
    row_num = len(listings) + 3
    for srow in _summary_rows(listings):
        if not srow:
            continue
        for col_idx, val in enumerate(srow, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if col_idx == 1:
                cell.font = Font(bold=True)
        row_num += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)


def export_pdf(listings: list[Listing], path: str | Path, title: str = "Listings") -> None:
    """Export listings to landscape A4 PDF."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, _strip_diacritics(title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    col_widths = [8, 20, 80, 25, 18, 22, 35, 22, 47]
    col_headers = ["#", "Source", "Title", "Price (EUR)", "Area", "EUR/m2", "Location", "Date", "URL"]

    pdf.set_font("Helvetica", "B", 7)
    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 5, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 6)
    for i, l in enumerate(listings, 1):
        pm2 = _price_per_m2(l)
        vals = [
            str(i),
            l.source[:3],
            l.title[:55],
            f"{l.price:,.0f}" if l.price else "",
            f"{l.area:.0f}" if l.area else "",
            f"{pm2:,.0f}" if pm2 else "",
            _location_str(l),
            l.date,
            (l.url or "")[:45],
        ]
        vals = [_strip_diacritics(v) for v in vals]
        for w, v in zip(col_widths, vals):
            pdf.cell(w, 4, v, border=1)
        pdf.ln()

    # Summary
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(0, 6, "Summary", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 8)

    prices = [l.price for l in listings if l.price is not None]
    pm2s = [v for v in (_price_per_m2(l) for l in listings) if v is not None]
    pdf.cell(0, 5, f"Listings with price: {len(prices)}  |  Listings with area: {len(pm2s)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    if prices:
        pdf.cell(0, 5, f"Price - Avg: {sum(prices)/len(prices):,.0f} EUR  |  Median: {median(prices):,.0f}  |  Min: {min(prices):,.0f}  |  Max: {max(prices):,.0f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    if pm2s:
        pdf.cell(0, 5, f"EUR/m2 - Avg: {sum(pm2s)/len(pm2s):,.0f}  |  Median: {median(pm2s):,.0f}  |  Min: {min(pm2s):,.0f}  |  Max: {max(pm2s):,.0f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.output(path)
