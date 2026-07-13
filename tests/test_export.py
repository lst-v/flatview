import csv

from flatview.export import (
    _location_str,
    _strip_diacritics,
    export_csv,
    export_pdf,
    export_xlsx,
)

# --- Helper function tests ---


def test_strip_diacritics():
    assert _strip_diacritics("Košice") == "Kosice"
    assert _strip_diacritics("Žilina") == "Zilina"
    assert _strip_diacritics("ASCII") == "ASCII"


def test_location_str_city_and_postcode(make_listing):
    l = make_listing(city="Michalovce", postcode="071 01")
    assert _location_str(l) == "Michalovce (071 01)"


def test_location_str_city_only(make_listing):
    l = make_listing(city="Bratislava", postcode="")
    assert _location_str(l) == "Bratislava"


# --- CSV export tests ---


def test_export_csv_creates_file(tmp_path, make_listing):
    path = tmp_path / "test.csv"
    listings = [make_listing(), make_listing(title="Second")]
    export_csv(listings, str(path))
    assert path.exists()


def test_export_csv_content(tmp_path, make_listing):
    path = tmp_path / "test.csv"
    l = make_listing(title="Test Byt", price=100000.0, area=50.0, source="bazos")
    export_csv([l], str(path))

    with open(path) as f:
        reader = csv.reader(f)
        rows = list(reader)

    assert rows[0] == [
        "#",
        "Source",
        "Segment",
        "Title",
        "Price (EUR)",
        "Area (m2)",
        "EUR/m2",
        "Location",
        "Date",
        "URL",
    ]
    assert rows[1][1] == "bazos"
    # rows[1][2] is the segment column (empty for unknown)
    assert rows[1][3] == "Test Byt"
    assert rows[1][4] == "100000.0"


def test_export_csv_summary_rows(tmp_path, make_listing):
    path = tmp_path / "test.csv"
    listings = [make_listing(price=100000.0), make_listing(price=200000.0)]
    export_csv(listings, str(path))

    with open(path) as f:
        content = f.read()

    assert "Average" in content
    assert "Median" in content
    assert "Min" in content
    assert "Max" in content
    assert "P25" in content
    assert "P75" in content


# --- XLSX export tests ---


def test_export_xlsx_creates_file(tmp_path, make_listing):
    path = tmp_path / "test.xlsx"
    export_xlsx([make_listing()], str(path))
    assert path.exists()
    assert path.stat().st_size > 0


def test_export_xlsx_header_bold(tmp_path, make_listing):
    from openpyxl import load_workbook

    path = tmp_path / "test.xlsx"
    export_xlsx([make_listing()], str(path))

    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(1, 1).font.bold


# --- PDF export tests ---


def test_export_pdf_creates_file(tmp_path, make_listing):
    path = tmp_path / "test.pdf"
    export_pdf([make_listing()], str(path))
    assert path.exists()
    assert path.stat().st_size > 0


def test_export_pdf_with_title(tmp_path, make_listing):
    path = tmp_path / "test.pdf"
    export_pdf([make_listing()], str(path), title="Michalovce 2izb")
    assert path.exists()
